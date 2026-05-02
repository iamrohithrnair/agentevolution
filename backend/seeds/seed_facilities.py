"""Seed ``facilities`` from the 9 hardcoded LOCATIONS + (optionally) the xlsx.

Run: ``uv run python -m backend.seeds.seed_facilities``

When ``backend/data/facilities.xlsx`` is present (489 rows from the original
DroneFleet repo) every row whose name does not collide with a hardcoded entry
is upserted with the same flat-earth ``airsim_xy`` projection. Without the
file the script seeds the 9 hardcoded sites only — the demo path needs them.
"""

from __future__ import annotations

import os
from pathlib import Path
from typing import Any

from motor.motor_asyncio import AsyncIOMotorDatabase
from pymongo import UpdateOne

from backend.dronan.bootstrap import apply_validator, FACILITIES_VALIDATOR

from ._common import bulk_upsert, latlon_to_xy, run, utcnow

# 9 canonical locations from the legacy ``config.py`` LOCATIONS dict.
HARDCODED_LOCATIONS: dict[str, dict[str, Any]] = {
    "Depot": {
        "x": 0, "y": 0, "z": -30, "lat": 51.5074, "lon": -0.1278,
        "description": "Main drone depot / base station", "type": "depot",
        "capabilities": [],
    },
    "Clinic A": {
        "x": 100, "y": 50, "z": -30, "lat": 51.5124, "lon": -0.1200,
        "description": "General medical clinic", "type": "clinic",
        "capabilities": ["urgent_care"],
    },
    "Clinic B": {
        "x": -50, "y": 150, "z": -30, "lat": 51.5174, "lon": -0.1350,
        "description": "Emergency care facility", "type": "clinic",
        "capabilities": ["urgent_care", "trauma"],
    },
    "Clinic C": {
        "x": 200, "y": -30, "z": -30, "lat": 51.5044, "lon": -0.1100,
        "description": "Rural health outpost", "type": "clinic",
        "capabilities": [],
    },
    "Clinic D": {
        "x": -100, "y": -80, "z": -30, "lat": 51.5000, "lon": -0.1400,
        "description": "Disaster relief camp", "type": "field_camp",
        "capabilities": ["urgent_care"],
    },
    "Royal London": {
        "x": 100, "y": 50, "z": -30, "lat": 51.5185, "lon": -0.0590,
        "description": "Royal London Hospital — Major trauma centre",
        "type": "hospital",
        "capabilities": ["trauma", "cardiac", "blood_bank", "cold_chain"],
    },
    "Homerton": {
        "x": -50, "y": 150, "z": -30, "lat": 51.5468, "lon": -0.0456,
        "description": "Homerton Hospital — Urgent care facility",
        "type": "hospital", "capabilities": ["urgent_care"],
    },
    "Newham General": {
        "x": 200, "y": -30, "z": -30, "lat": 51.5155, "lon": 0.0285,
        "description": "Newham General Hospital — Trauma kit resupply",
        "type": "hospital", "capabilities": ["trauma"],
    },
    "Whipps Cross": {
        "x": -100, "y": -80, "z": -30, "lat": 51.5690, "lon": 0.0066,
        "description": "Whipps Cross Hospital — Cardiac unit",
        "type": "hospital", "capabilities": ["cardiac"],
    },
}

DEPOT_LAT = HARDCODED_LOCATIONS["Depot"]["lat"]
DEPOT_LON = HARDCODED_LOCATIONS["Depot"]["lon"]

XLSX_PATH = Path(
    os.environ.get(
        "DRONAN_FACILITIES_XLSX",
        str(Path(__file__).resolve().parent.parent / "data" / "facilities.xlsx"),
    )
)


def _normalize_type(t: str | None) -> str:
    t = (t or "").strip().lower()
    if "hospital" in t:
        return "hospital"
    if "clinic" in t:
        return "clinic"
    if "depot" in t or "warehouse" in t:
        return "depot"
    if "camp" in t or "field" in t:
        return "field_camp"
    return "clinic"


def _to_doc(
    name: str,
    *,
    lat: float,
    lon: float,
    type_: str,
    address: str | None = None,
    region: str | None = None,
    phone: str | None = None,
    email: str | None = None,
    website: str | None = None,
    capabilities: list[str] | None = None,
    description: str | None = None,
    source: str = "xlsx",
) -> dict[str, Any]:
    x, y = latlon_to_xy(lat, lon, DEPOT_LAT, DEPOT_LON)
    return {
        "name": name,
        "type": type_,
        "region": region,
        "address": address,
        "phone": phone,
        "email": email,
        "website": website,
        "capabilities": list(capabilities or []),
        "location": {"type": "Point", "coordinates": [float(lon), float(lat)]},
        "airsim_xy": {"x": float(x), "y": float(y), "z": -30.0},
        "description": description,
        "source": source,
        "created_at": utcnow(),
        "updated_at": utcnow(),
    }


def _load_xlsx_rows(path: Path) -> list[dict[str, Any]]:
    if not path.is_file():
        return []
    try:
        from openpyxl import load_workbook
    except ImportError:  # pragma: no cover
        return []

    wb = load_workbook(path, read_only=True)
    ws = wb.active
    headers: list[str] | None = None
    rows: list[dict[str, Any]] = []
    for row in ws.iter_rows(values_only=True):
        if headers is None:
            headers = [
                str(h).strip().lower().replace(" ", "_") if h else "" for h in row
            ]
            continue
        rec = dict(zip(headers, row))
        try:
            lat = float(rec.get("latitude") or 0)
            lon = float(rec.get("longitude") or 0)
        except (TypeError, ValueError):
            continue
        if lat == 0 and lon == 0:
            continue
        name = (rec.get("name") or "").strip()
        if not name:
            continue
        rows.append(
            {
                "name": name,
                "type": _normalize_type(rec.get("type")),
                "region": (rec.get("region") or "").strip() or None,
                "address": (rec.get("physical_address") or "").strip() or None,
                "phone": (rec.get("phone_number") or "").strip() or None,
                "email": (rec.get("email") or "").strip() or None,
                "website": (rec.get("website") or "").strip() or None,
                "lat": lat,
                "lon": lon,
            }
        )
    wb.close()
    return rows


async def main(db: AsyncIOMotorDatabase) -> dict[str, int]:
    """Idempotently upsert facilities. Returns counts for verification."""
    await apply_validator(db, "facilities", FACILITIES_VALIDATOR)

    ops: list[UpdateOne] = []

    # 1. Hardcoded 9
    for name, c in HARDCODED_LOCATIONS.items():
        doc = _to_doc(
            name,
            lat=c["lat"],
            lon=c["lon"],
            type_=c.get("type", "clinic"),
            description=c.get("description"),
            capabilities=list(c.get("capabilities", [])),
            source="config",
        )
        # Preserve original AirSim x/y from the hardcoded values.
        doc["airsim_xy"] = {"x": float(c["x"]), "y": float(c["y"]), "z": float(c["z"])}
        ops.append(UpdateOne({"name": name}, {"$set": doc}, upsert=True))

    # 2. xlsx (skip rows whose name collides with a hardcoded one)
    rows = _load_xlsx_rows(XLSX_PATH)
    for r in rows:
        if r["name"] in HARDCODED_LOCATIONS:
            continue
        doc = _to_doc(
            r["name"],
            lat=r["lat"],
            lon=r["lon"],
            type_=r["type"],
            address=r["address"],
            region=r["region"],
            phone=r["phone"],
            email=r["email"],
            website=r["website"],
            source="xlsx",
        )
        ops.append(UpdateOne({"name": r["name"]}, {"$set": doc}, upsert=True))

    if not ops:
        return {"upserted": 0, "modified": 0, "total": 0}

    res = await bulk_upsert(db.facilities, ops)
    total = await db.facilities.count_documents({})
    print(
        f"facilities: upserted={res['upserted']} "
        f"modified={res['modified']} total={total}"
    )
    return {**res, "total": total}


if __name__ == "__main__":
    run(main)
